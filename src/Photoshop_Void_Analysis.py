from pathlib import Path
import cv2
import numpy as np
import openpyxl
from scipy.ndimage import label, center_of_mass

def calculate_void_percent(full_image, cut=False):
    """
    Calculate the percentage of pixels in the image that are red.
    """

    if cut is not False:
        # splice up the image
        cut_length = round(max(full_image.shape) * cut)
        image = full_image[cut_length:full_image.shape[1] - cut_length, cut_length:full_image.shape[0] - cut_length]

    else:
        image = full_image
    # Convert the image from BGR to HSV color space
    hsv_image = cv2.cvtColor(full_image, cv2.COLOR_BGR2HSV)

    # Define the lower and upper bounds for red in HSV
    # Red can span over two ranges in the HSV space
    lower_red1 = np.array([0, 50, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 50, 50])
    upper_red2 = np.array([180, 255, 255])

    # Create masks for the two red ranges
    mask1 = cv2.inRange(hsv_image, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv_image, lower_red2, upper_red2)

    # Combine the masks
    red_mask = cv2.bitwise_or(mask1, mask2)

    # Count the number of red pixels
    red_pixel_count = np.sum(red_mask > 0)

    # Calculate the total number of pixels in the image
    total_pixels = image.shape[0] * image.shape[1]

    # Calculate the percentage of red pixels
    percentage_red = float((red_pixel_count / total_pixels) * 100)

    return percentage_red

def load_images(image_identifier):

    script_dir = Path(__file__).parent
    pre_cycle_dir = script_dir.parent / 'Non-code' / 'Void Study pre-cycle'
    post_cycle_dir = script_dir.parent / 'Non-code' / 'Void Study post-cycle'

    pre_images_full = dict()
    post_images_full = dict()
    
    # load the pre-cycle images into the dict
    for label_dir in pre_cycle_dir.rglob("*"):
        if label_dir.is_dir():

            for file_path in label_dir.rglob("*"):
                file_str = str(file_path).split("\\")
                filename = file_str[len(file_str)-1].split('.')[0]

                if filename[len(filename)-len(image_identifier) :] == image_identifier:
                    label = filename.replace(image_identifier, "")
                    pre_images_full[label] = cv2.imread(file_path)


    # load the post-cycle images into the dict
    for label_dir in post_cycle_dir.rglob("*"):
        if label_dir.is_dir():

            for file_path in label_dir.rglob("*"):
                file_str = str(file_path).split("\\")
                filename = file_str[len(file_str)-1].split('.')[0]
                
                if filename[len(filename)-len(image_identifier) :] == image_identifier:
                    label = filename.replace('vP', "")
                    post_images_full[label] = cv2.imread(file_path)

    return pre_images_full, post_images_full

def void_all_images(pre_images, post_images, cut=False):

    pre_voids = dict()
    post_voids = dict()

    for label in pre_images.keys():
        pre_voids[label] = calculate_void_percent(pre_images[label], cut)

    for label in post_images.keys():
        post_voids[label] = calculate_void_percent(post_images[label], cut)

    return pre_voids, post_voids

def get_void_sizes_and_positions(full_image, cut=False):
    """
    This will return a list of tuples, with the pixel coordinates of the centroid of the void as the first term and the ratio of the void size to the image size as the second.
    """
    if cut is not False:
        # splice up the image
        cut_length = round(max(full_image.shape) * cut)
        image = full_image[cut_length:full_image.shape[1] - cut_length, cut_length:full_image.shape[0] - cut_length]

    else:
        image = full_image

    # Convert the image from BGR to HSV color space
    hsv_image = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)

    # Define the lower and upper bounds for red in HSV
    # Red can span over two ranges in the HSV space
    lower_red1 = np.array([0, 50, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 50, 50])
    upper_red2 = np.array([180, 255, 255])

    # Create masks for the two red ranges
    mask1 = cv2.inRange(hsv_image, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv_image, lower_red2, upper_red2)

    # Combine the masks
    red_mask = cv2.bitwise_or(mask1, mask2)

    # Find connected components (blobs)
    labeled_mask, num_voids = label(red_mask)
    
    # Extract the positions and sizes of the blobs
    void_positions_and_sizes = []
    for i in range(1, num_voids + 1):
        centroid = center_of_mass(red_mask, labeled_mask, i)
        centroid = (int(centroid[1]), int(centroid[0]))

        size = int(np.sum(labeled_mask == i))  # Area in pixels
        relative_size = 100 * size / (image.shape[0] * image.shape[1])
        void_positions_and_sizes.append((centroid, relative_size))

        #cv2.circle(image, centroid, radius=1, color=(255, 0, 0), thickness=-1)  # Blue dot
    
    #cv2.imshow("Centroids Marked", image)
    #cv2.waitKey(0)
    #cv2.destroyAllWindows()

    return void_positions_and_sizes

def main():
    """
    Go to the photoshopped images, and calculate the void percentage.
    """

    image_identifier = 'vP'
    pre_images_full, post_images_full = load_images(image_identifier)
    
    # calculate voids for all images
    pre_voids, post_voids = void_all_images(pre_images_full, post_images_full, cut=False)
    print(pre_voids)

    writeToExcel = False
    if writeToExcel is True:
        # Write the results to the excel file
        script_dir = Path(__file__).parent
        excel_file_path = script_dir.parent / 'Experimental Data' / 'Void Study FULL DOC.xlsx'
        excel_sheet = openpyxl.load_workbook(excel_file_path)
        workbook_name = 'Photoshop Void Data (2)'

        excel_sheet.create_sheet(title=workbook_name)
        for i, label in enumerate(pre_voids.keys()):
            excel_sheet[f'{workbook_name}'].cell(row=1, column=i+1).value = label
            excel_sheet[f'{workbook_name}'].cell(row=2, column=i+1).value = pre_voids[label]

        excel_sheet.create_sheet(title=f'Post-Cycle {workbook_name}')
        for i, label in enumerate(post_voids.keys()):
            excel_sheet[f'Post-Cycle {workbook_name}'].cell(row=1, column=i+1).value = label
            excel_sheet[f'Post-Cycle {workbook_name}'].cell(row=2, column=i+1).value = post_voids[label]

        excel_sheet.save(excel_file_path)

main()